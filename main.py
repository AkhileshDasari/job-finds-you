"""
Run all company scrapers, filter for entry-level/intern IT roles, and export
results to CSV + a nicely formatted Excel file.

Usage:
    python main.py
    python main.py --with-salary       # slower: opens each Workday job page to look for stipend/salary text
    python main.py --output-dir ./out

Output files are written to ./output/ by default:
    jobs_YYYY-MM-DD.csv
    jobs_YYYY-MM-DD.xlsx
"""
import argparse
import datetime
import os
import sys

import requests
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import COMPANIES
from keywords import SEARCH_TERMS, is_relevant, classify_type, extract_stipend, is_location_relevant
from scrapers.workday import fetch_workday_jobs, fetch_workday_job_description
from scrapers.json_apis import fetch_google_jobs, fetch_microsoft_jobs, fetch_apple_jobs, fetch_amazon_jobs
from scrapers.fallback import build_fallback_rows

REAL_SCRAPE_METHODS = {"workday", "google", "microsoft", "apple", "amazon"}


def scrape_company(company, session, with_salary=False):
    """Returns a list of row-dicts for one company."""
    method = company["method"]
    name = company["name"]
    print(f"  -> {name} ({method})")

    if method == "workday":
        jobs = fetch_workday_jobs(company["tenant"], company["wd_host"], company["site"],
                                   SEARCH_TERMS, session=session)
        rows = []
        for j in jobs:
            if not is_relevant(j["title"]):
                continue
            if not is_location_relevant(j.get("location", "")):
                continue
            stipend = "Not specified"
            if with_salary and j.get("external_path"):
                desc = fetch_workday_job_description(company["tenant"], company["wd_host"],
                                                       company["site"], j["external_path"], session=session)
                stipend = extract_stipend(desc)
            rows.append({
                "Company": name, "Category": company["category"],
                "Role Title": j["title"], "Type": classify_type(j["title"]),
                "Location": j.get("location", "Not specified"),
                "Apply Link": j["link"], "Stipend / Salary": stipend,
                "Posted": j.get("posted", ""),
            })
        return rows

    if method in ("google", "microsoft", "apple", "amazon"):
        fetch_fn = {"google": fetch_google_jobs, "microsoft": fetch_microsoft_jobs,
                    "apple": fetch_apple_jobs, "amazon": fetch_amazon_jobs}[method]
        jobs = fetch_fn(SEARCH_TERMS, session=session)
        rows = []
        for j in jobs:
            if not is_relevant(j["title"]):
                continue
            if not is_location_relevant(j.get("location", "")):
                continue
            rows.append({
                "Company": name, "Category": company["category"],
                "Role Title": j["title"], "Type": classify_type(j["title"]),
                "Location": j.get("location", "Not specified"),
                "Apply Link": j["link"], "Stipend / Salary": "Not specified",
                "Posted": j.get("posted", ""),
            })
        return rows

    if method == "fallback":
        jobs = build_fallback_rows(company["url"], SEARCH_TERMS)
        return [{
            "Company": name, "Category": company["category"],
            "Role Title": j["title"], "Type": "Manual check required",
            "Location": j.get("location", "N/A"),
            "Apply Link": j["link"], "Stipend / Salary": "Not specified",
            "Posted": "",
        } for j in jobs]

    raise ValueError(f"Unknown method: {method}")


def export_excel(df, path):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Listings")
        ws = writer.sheets["Listings"]

        header_fill = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
        header_font = Font(bold=True, color="FFFFFF", name="Arial")
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        widths = {"Company": 28, "Category": 10, "Role Title": 50, "Type": 22,
                  "Location": 30, "Apply Link": 55, "Stipend / Salary": 22, "Posted": 14}
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 18)

        link_col = list(df.columns).index("Apply Link") + 1
        for row_idx in range(2, len(df) + 2):
            cell = ws.cell(row=row_idx, column=link_col)
            if cell.value:
                cell.hyperlink = cell.value
                cell.font = Font(color="0563C1", underline="single", name="Arial")
            ws.cell(row=row_idx, column=1).font = Font(name="Arial")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions


def main():
    parser = argparse.ArgumentParser(description="Scrape entry-level/intern IT roles from major MNCs.")
    parser.add_argument("--with-salary", action="store_true",
                         help="Also open each Workday job page to look for stipend/salary text (slower).")
    parser.add_argument("--output-dir", default="output")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    session = requests.Session()

    all_rows = []
    print(f"Scraping {len(COMPANIES)} companies...\n")
    for company in COMPANIES:
        try:
            rows = scrape_company(company, session, with_salary=args.with_salary)
        except Exception as exc:
            print(f"    [error] {company['name']} failed entirely: {exc}")
            rows = []
        real = company["method"] in REAL_SCRAPE_METHODS
        tag = f"{len(rows)} matching roles found" if real else f"{len(rows)} search link(s) generated"
        print(f"     {tag}")
        all_rows.extend(rows)

    if not all_rows:
        print("\nNo rows collected. Check your internet connection / company site availability.")
        sys.exit(1)

    df = pd.DataFrame(all_rows, columns=["Company", "Category", "Role Title", "Type",
                                          "Location", "Apply Link", "Stipend / Salary", "Posted"])
    df = df.sort_values(["Category", "Company", "Type"]).reset_index(drop=True)

    today = datetime.date.today().isoformat()
    csv_path = os.path.join(args.output_dir, f"jobs_{today}.csv")
    xlsx_path = os.path.join(args.output_dir, f"jobs_{today}.xlsx")

    df.to_csv(csv_path, index=False)
    export_excel(df, xlsx_path)

    n_real = (df["Type"] != "Manual check required").sum()
    n_manual = (df["Type"] == "Manual check required").sum()
    print(f"\nDone. {n_real} scraped roles + {n_manual} manual-check links -> {len(df)} total rows.")
    print(f"  CSV : {os.path.abspath(csv_path)}")
    print(f"  XLSX: {os.path.abspath(xlsx_path)}")


if __name__ == "__main__":
    main()
